from openpyxl import load_workbook


def get_departments():
    wb = load_workbook('sample/departments.xlsx')

    departments = []
    for row in wb.active.iter_rows(min_row=2):
        dep = {
            'externalId': row[0].value,
            'names': {
            }
        }
        if row[1].value:
            dep['names']['fi'] = row[1].value
        if row[2].value:
            dep['names']['sv'] = row[2].value
        if row[3].value:
            dep['names']['en'] = row[3].value

        departments.append(dep)

    return departments


def get_personnel():
    wb = load_workbook('sample/personnel.xlsx')

    workers = []
    for row in wb.active.iter_rows(min_row=2):
        employee = {
            'externalId': row[0].value,
            'ssn': row[3].value,
            'callName': row[1].value,
            'lastName': row[2].value,
            'emailAddress': row[4].value,
            'localPhoneNumber': row[7].value,
            'startDate': row[8].value.strftime('%Y-%m-%d'),
            'departments': [{
                'externalId': row[23].value,
                'startDate': row[8].value.strftime('%Y-%m-%d'),
            }]
        }
        if row[9].value:
            employee['endDate'] = row[9].value.strftime('%Y-%m-%d')
        if row[18].value and row[20].value:
            employee['supervisors'] = [{
                'externalId': row[18].value,
                'startDate': row[20].value.strftime('%Y-%m-%d')
            }]
        workers.append(employee)

    return workers
