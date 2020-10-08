from openpyxl import load_workbook
import datetime


def get_absences():
    wb = load_workbook('sample/absences.xlsx')
    absences = []
    for row in wb.active.iter_rows(min_row=2):
        absence = {
            'externalId': row[0].value,
            'startDate': row[1].value.strftime('%Y-%m-%d')
        }
        if row[2].value:
            absence['endDate'] = row[2].value.strftime('%Y-%m-%d')
        absences.append(absence)

    return absences
