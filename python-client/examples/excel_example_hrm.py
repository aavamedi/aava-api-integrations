import utils
from openpyxl import load_workbook


def get_departments():
    # Load properties and check that they are structurally OK
    props = utils.load_properties()
    try:
        assert props['hrMgmtSystem'] != None, 'No HRM properties section'
        assert props['hrMgmtSystem']['departmentsFile'] != None, 'HRM Departments Excel file not set'
    except Exception as ex:
        print("Properties file not complete:", repr(ex))
        exit()

    wb = load_workbook(props['hrMgmtSystem']['departmentsFile'])

    departments = []
    for row in wb.active.iter_rows(min_row=2):
        external_id, fi, sv, en = row
        dep = {
            'externalId': external_id.value,
            'names': {
            }
        }
        if row[1].value:
            dep['names']['fi'] = fi.value
        if row[2].value:
            dep['names']['sv'] = sv.value
        if row[3].value:
            dep['names']['en'] = en.value

        departments.append(dep)

    return departments


def get_get_cost_centers():
    # this function intentionally returns nothing, structurally it
    # can resemble 'get_departments'
    return []


def get_personnel():
    # Load properties and check that they are structurally OK
    props = utils.load_properties()
    try:
        assert props['hrMgmtSystem'] != None, 'No HRM properties section'
        assert props['hrMgmtSystem']['employeeFile'] != None, 'HRM Employees Excel file not set'
    except Exception as ex:
        print("Properties file not complete:", repr(ex))
        exit()

    wb = load_workbook(props['hrMgmtSystem']['employeeFile'])

    employees = []
    for row in wb.active.iter_rows(min_row=2):
        external_id, identifier, ssn, call_name, last_name, email_address, private_email_address, \
            job_title, local_phone_number, phone_country_code, start_date, end_date, department, department_start, \
            supervisor, supervisor_start = row

        employee = {
            'externalId': external_id.value,
            'identifier': identifier.value,
            'ssn': ssn.value,
            'callName': call_name.value,
            'lastName': last_name.value,
            'emailAddress': email_address.value,
            'privateEmailAddress': private_email_address.value,
            'jobTitle': job_title.value,
            'localPhoneNumber': local_phone_number.value,
            'phoneCountryCode': phone_country_code.value,
            'startDate': start_date.value.strftime('%Y-%m-%d'),
            'departments': [{
                'externalId': department.value,
                'startDate': department_start.value.strftime('%Y-%m-%d'),
            }]
        }
        if end_date.value:
            employee['endDate'] = end_date.value.strftime('%Y-%m-%d')
        if supervisor.value and supervisor_start.value:
            employee['supervisors'] = [{
                'externalId': supervisor.value,
                'startDate': supervisor_start.value.strftime('%Y-%m-%d')
            }]
        employees.append(employee)

    return employees
